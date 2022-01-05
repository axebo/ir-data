import paramiko
import os
import csv
from decimal import Decimal
from scp import SCPClient
from datetime import datetime, timedelta
from openpyxl import load_workbook

def createSSHClient(server, user, keyfile):
    client = paramiko.SSHClient()
    client.load_system_host_keys()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(server, username=user, key_filename=keyfile)
    return client

def processLog(logFile, excelFile):
    csvOut = []
    with open(logFile) as f:
        csvReader = csv.reader(f)
        for row in csvReader:
            t = row[0][:-1]
            if len(t) > 26:
                t = t[:26]
            time = datetime.strptime(t,'%Y-%m-%dT%H:%M:%S.%f')
            csvOut.append([time, Decimal(row[1]), Decimal(row[2])])
    wb = load_workbook(filename=excelFile)
    for row in csvOut:
        wb['AllTrades'].insert_rows(2, amount=1)
        for col, val in enumerate(row, start=1):
            wb['AllTrades'].cell(row=2, column=col).value = val

    wb.save(filename=excelFile)

ssh = createSSHClient('159.89.205.244', 'irbot', 'C:\\Users\\base\\biit.pem')
scp = SCPClient(ssh.get_transport())
logDir = r'C:\Users\base\Documents\Endeavours\Crypt\ir\IRtradeslog'

logList = os.listdir(logDir)
print(logList[-1])
latestLogDate = datetime.strptime(sorted(logList)[-1][:8],'%Y%m%d').date()

#logs in UTC date, can get log after next day 11am sydney time
LatestAvailLogDate = (datetime.now()+timedelta(days=-1,hours=-10)).date()
while latestLogDate < LatestAvailLogDate:
    latestLogDate += timedelta(days=1)
    logName = latestLogDate.strftime('%Y%m%d')+'_irrecords.csv'
    print('looking for:', logName)
    try:
        scp.get('/home/irbot/indep/log/'+logName, logDir)
        print(f'''received {latestLogDate.strftime('%Y%m%d')}''')
        processLog(logDir+'\\'+logName, r'C:\Users\base\Documents\Endeavours\Crypt\ir\IRtrades.xlsx')
        print(f'''processed {latestLogDate.strftime('%Y%m%d')}''')
    except:
        print(f'''problem with {latestLogDate}''')
